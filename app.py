from openpyxl import load_workbook, Workbook
from flask import Flask, render_template, request, redirect, url_for
from flask_sqlalchemy import SQLAlchemy
import pandas as pd
import random

# Read the Excel file

id1 = 53
id2 = 52
app = Flask(__name__)
username = 'postgres'
password = 'user'
host = 'localhost'
port = '5432'
database_name = 'carcmp'

database_uri = f'postgresql://{username}:{password}@{host}:{port}/{database_name}'


app.config['SQLALCHEMY_DATABASE_URI'] = database_uri
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)




class datas(db.Model):
    __tablename__ = 'datas'
    id = db.Column(db.Integer,primary_key = True)
    brand = db.Column(db.String)
    model = db.Column(db.String)
    year = db.Column(db.String)
    price = db.Column(db.String)
    type = db.Column(db.String)
    colours = db.Column(db.String)
    fuel = db.Column(db.String)
    mileage = db.Column(db.String)
    boot_space = db.Column(db.String)
    seat_capacity = db.Column(db.String)
    tyre_size = db.Column(db.String)
    air_bags = db.Column(db.String)
    cruise_ctrl = db.Column(db.String)
    engine = db.Column(db.String)
    cylinders = db.Column(db.String)
    transmission = db.Column(db.String)
    gear_box = db.Column(db.String)
    drive_type = db.Column(db.String)
    description = db.Column(db.String)
    recommentation = db.Column(db.String)

        


    
def create_excel_file_if_not_exists():
    try:
        wb = load_workbook('form_responses.xlsx')
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(['First Name', 'Last Name', 'Email', 'Message'])
        wb.save('form_responses.xlsx')

class kart(db.Model):
    __tablename__ = 'Kart'
    brand = db.Column(db.String)
    carsid = db.Column(db.Integer)
    id = db.Column(db.Integer, primary_key=True)    
    model =db.Column(db.String)
    description = db.Column(db.String)
    price = db.Column(db.String)

class Wishlist(db.Model):
    __tablename__ = 'wishlist'
    brand = db.Column(db.String)
    carsid = db.Column(db.Integer)
    id = db.Column(db.Integer, primary_key=True)    
    model =db.Column(db.String)
    description = db.Column(db.String)
    price = db.Column(db.String)

class cars(db.Model):
    __tablename__ = 'carcmp'
    id = db.Column(db.Integer ,primary_key = True)

create_excel_file_if_not_exists()

@app.route('/')
def index():
    unique_brands = db.session.query(datas.brand).distinct().all()
    unique_seats = db.session.query(datas.seat_capacity).distinct().all()
    unique_colors = db.session.query(datas.colours).distinct().all()
    unique_fuel = db.session.query(datas.fuel).distinct().all()
    row = datas.query.all()
    return render_template('index.html',rows= row,fun = 'all',brands = unique_brands,colours = unique_colors,fuels = unique_fuel,seats = unique_seats)

   

@app.route('/importing')
def excel_file_import():
    excel_file = 'cars.xlsx'
    df = pd.read_excel(excel_file)
    for index, row in df.iterrows():
        mileage = str(row['mileage']).split()[0]
        price = str(row['price']).replace(' ','')
        boot_spac=str(row['bootspace']).split()[0]
        seat_capacity = str(row['seating_capacity']).split()[0]
        # Create a new datas object
        data = datas(
            brand=row['brand'].upper(),
            model=row['model'].upper(),
            year=row['year'],
            price=price,
            type=str(row['type']).split()[0],
            colours=str(row['colors']).split()[0],
            fuel=row['feul'],
            mileage=mileage,
            boot_space=boot_spac,
            seat_capacity=seat_capacity,
            tyre_size=str(row['tyresize']).replace(' ',''),
            air_bags=str(row['airbags']),
            cruise_ctrl=str(row['cruisectrl']),
            engine=str(row['engine']).split()[0],
            cylinders=str(row['cylinders']).split()[0],
            transmission=str(row['transmission ']).split()[0],
            gear_box=str(row['gearbox']).split()[0],
            drive_type=str(row['drivetype']),
            description=str(row['description']),
            recommentation= str(random.randint(80, 100))
        )
        
        db.session.add(data)
    db.session.commit()
    return "Commit completed"


@app.route('/contact')
def contactpage():
    return render_template('contactus.html')

@app.route('/preview/<id>')
def preview(id):
    row = datas.query.get(id)
    return render_template('preview.html',row=row)

@app.route('/addtocart/<id>')
def addcart(id):
    row = datas.query.get(id)
    data = kart( carsid = id,model = row.model ,brand = row.brand, description = row.description,price = row.price)
    db.session.add(data)
    db.session.commit()
    return redirect(url_for('cart'))

@app.route('/remove/<id>')
def remove(id):
    row = Wishlist.query.get(id)
    if row:
        db.session.delete(row)
        db.session.commit()
    return redirect(url_for('wishlist'))


@app.route('/addtowishlist/<id>')
def addwishlist(id):
    row = datas.query.get(id)
    data = Wishlist( carsid = id,model = row.model ,brand = row.brand, description = row.description,price = row.price)
    db.session.add(data)
    db.session.commit()
    return redirect(url_for('wishlist'))


@app.route('/addcompare/<id>')
def addcompare(id):
    global id2,id1
    id2= id1
    id1 = id
    print(id1,id2)
    return redirect(url_for('compare'))

@app.route('/filter', methods=['POST', 'GET'])
def filter():
    if request.method == 'POST':
        fuel = ""
        color = ""

        seat = request.form.get('seating')  # Use request.form.get to get form data
        brand = request.form.get('brand')
        color = request.form.get('color')
        if color == None:
            color = ''
        fuel = request.form.get('fuel')
        query = db.session.query(datas)
        if seat != "":
            query = query.filter(datas.seat_capacity == seat)  # Correct variable name to seat
        if brand != "":
            query = query.filter(datas.brand == brand)
        if color != "":
            query = query.filter(datas.colours == color)
        if fuel != "":
            query = query.filter(datas.fuel == fuel)

        print(seat,fuel,color,brand)
        unique_brands = db.session.query(datas.brand).distinct().all()
        unique_seats = db.session.query(datas.seat_capacity).distinct().all()
        unique_colors = db.session.query(datas.colours).distinct().all()
        unique_fuel = db.session.query(datas.fuel).distinct().all()
        results = query.all()
        return render_template('allcarpreview.html', rows=results, fun='all',brands = unique_brands,colours = unique_colors,fuels = unique_fuel,seats = unique_seats)


@app.route('/search',methods = ['POST','GET'])
def search():
    if request.method == 'POST':
        search = request.form.get('searchname').upper()
        query = db.session.query(datas)
        carname = query.filter(datas.brand == search or datas.model == search)
    return render_template('allcarpreview.html', rows = carname, fun = 'all')


@app.route('/compare')
def compare():
    global id2,id1
    row1 = datas.query.get(id1)
    row2 = datas.query.get(id2)
    return render_template('comparison.html',row1 = row1 ,row2 = row2)

@app.route('/contactus', methods=['POST', 'GET'])
def contact():
    if request.method == 'POST':
        fname = request.form.get('fname')
        lname = request.form.get('lname')
        email = request.form.get('email')
        message = request.form.get('message')
        wb = load_workbook('form_responses.xlsx')
        ws = wb.active
        ws.append([fname, lname, email, message])
        wb.save('form_responses.xlsx')

        return "Form responses saved successfully!"

    return "Method Not Allowed"


@app.route('/audicars')
def audi():
    row = datas.query.filter(datas.brand == 'audi').all()
    return render_template('allcarpreview.html',rows= row,fun = 'all')


@app.route('/volvocars')
def volvo():
    row = datas.query.filter(datas.brand == 'volvo').all()
    return render_template('allcarpreview.html',rows= row,fun = 'all')


@app.route('/cartview')
def cart():
    row = kart.query.all()
    return render_template('allcarpreview.html',rows= row,func = 'cart')



@app.route('/wishlistview')
def wishlist():
    row = Wishlist.query.all()
    return render_template('allcarpreview.html',rows= row,func = 'cart')

@app.route('/allcars')
def all():
    unique_brands = db.session.query(datas.brand).distinct().all()
    unique_seats = db.session.query(datas.seat_capacity).distinct().all()
    unique_colors = db.session.query(datas.colours).distinct().all()
    unique_fuel = db.session.query(datas.fuel).distinct().all()
    row = datas.query.all()
    return render_template('allcarpreview.html',rows= row,fun = 'all',brands = unique_brands,colours = unique_colors,fuels = unique_fuel,seats = unique_seats)


@app.route('/hatchback')
def hatchback():
    row = datas.query.filter(datas.type == 'hatchback').all()
    unique_brands = db.session.query(datas.brand).distinct().all()
    unique_seats = db.session.query(datas.seat_capacity).distinct().all()
    unique_colors = db.session.query(datas.colours).distinct().all()
    unique_fuel = db.session.query(datas.fuel).distinct().all()
    return render_template('allcarpreview.html',rows= row,fun = 'all',brands = unique_brands,colours = unique_colors,fuels = unique_fuel,seats = unique_seats)

@app.route('/sedan')
def sedan():
    row = datas.query.filter(datas.type == 'sedan').all()
    unique_brands = db.session.query(datas.brand).distinct().all()
    unique_seats = db.session.query(datas.seat_capacity).distinct().all()
    unique_colors = db.session.query(datas.colours).distinct().all()
    unique_fuel = db.session.query(datas.fuel).distinct().all()
    return render_template('allcarpreview.html',rows= row,fun = 'all',brands = unique_brands,colours = unique_colors,fuels = unique_fuel,seats = unique_seats)




if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=False) 

